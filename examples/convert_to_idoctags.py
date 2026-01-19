import os
import json
import glob
import argparse
from pathlib import Path
from typing import Sequence, Dict, Any, Optional
from collections import Counter
from io import BytesIO

from PIL import Image as PILImage

from datasets import load_dataset
from tqdm import tqdm
from transformers import (
    AutoTokenizer,
    PreTrainedTokenizerBase,
)
from docling_core.types.doc import DoclingDocument, ImageRef
from docling_core.types.doc.base import ImageRefMode
from docling_core.experimental.idoctags import (
    EscapeMode,
    IDocTagsSerializationMode,
    IDocTagsParams,
    IDocTagsVocabulary,
    IDocTagsDocSerializer,
)

import matplotlib.pyplot as plt
import numpy as np

# In order to download **before** the datasets library, run
#
# HF_HUB_DISABLE_XET=1 hf download --repo-type dataset "{hf-repo-id}"
#

def update_tokenizer(tokenizer: PreTrainedTokenizerBase, verbose: bool = False) -> PreTrainedTokenizerBase:
    """Extend tokenizer with IDocTags special tokens.

    Parameters
    - tokenizer: base tokenizer to extend
    - verbose: print added tokens if True
    """
    special_tokens = IDocTagsVocabulary.get_special_tokens()
    if verbose:
        for i, tok in enumerate(special_tokens):
            print(i, "\t", tok)
    tokenizer.add_tokens(special_tokens)
    if verbose:
        print(f"New vocab size: {tokenizer.vocab_size}")
    return tokenizer

def run_dump(cfg: dict[str, Any]) -> int:
    """Dump/serialize documents from a dataset to IDocTags strings/files and export a per-row report.

    Config keys (with defaults):
    - dataset_name: str ("docling-project/doclaynet-set-a")
    - dataset_subset: str ("pdf_train")
    - dataset_split: str ("train")
    - output_dir: str ("./scratch/idoctags")
    - failed_dir: str ("./scratch/idoctags_failed") — where to dump HTML+JSON when serialization fails
    - write_outputs: bool (True) — write serialized outputs if True
    - report_path: str ("./scratch/idoctags_report.xlsx") — where to write the results table (xlsx or csv)
    - limit: Optional[int] (None) — process only the first N items if set
    - variants: list of serialization variants; each item:
        {"add_content": bool, "mode": "LLM_FRIENDLY"|"HUMAN_FRIENDLY", "suffix": str}
      Defaults to three variants mirroring prior behavior.
    """
    dataset_name = cfg.get("dataset_name", "docling-project/doclaynet-set-a")
    dataset_subset = cfg.get("dataset_subset", "train")
    dataset_split = cfg.get("dataset_split", "train")
    output_dir = Path(cfg.get("output_dir", "./scratch/idoctags"))
    failed_dir = Path(cfg.get("failed_dir", "./scratch/idoctags_failed"))
    pngs_dir = Path(cfg.get("pngs_dir", "./scratch/pngs_dir"))
    write_outputs: bool = bool(cfg.get("write_outputs", True))
    report_path = Path(cfg.get("report_path", "./scratch/idoctags_report.xlsx"))

    default_variants = [
        {"add_content": False, "mode": "LLM_FRIENDLY", "suffix": "_without"},
        {"add_content": True, "mode": "LLM_FRIENDLY", "suffix": "_with"},
        {"add_content": True, "mode": "HUMAN_FRIENDLY", "suffix": "_with_h"},
    ]
    variants = cfg.get("variants", default_variants)

    if write_outputs:
        output_dir.mkdir(parents=True, exist_ok=True)
        # Create failed dir proactively; also created again on demand in the exception path
        failed_dir.mkdir(parents=True, exist_ok=True)

    ds = load_dataset(dataset_name, dataset_subset)
    split = ds[dataset_split]
    total = len(split)

    # Effective processing limit
    raw_limit = cfg.get("limit") or cfg.get("max_items") or cfg.get("max_samples")
    limit: Optional[int] = None
    if isinstance(raw_limit, int) and raw_limit > 0:
        limit = raw_limit
    processed_total = min(total, limit) if limit else total

    errors: list[str] = []

    # Results rows for report
    results_rows: list[dict[str, str]] = []

    def _yes(b: bool) -> str:
        return "Yes" if b else "No"

    def _sanitize_cell_value(value: str, warn: bool = False) -> str:
        """Sanitize cell value to ensure XML/Excel compliance.

        Removes or replaces characters that are invalid in XML, which would
        cause openpyxl to fail with 'not well-formed (invalid token)' errors.

        Invalid XML characters include most control characters (0x00-0x1F)
        except tab (0x09), newline (0x0A), and carriage return (0x0D).

        Parameters
        - value: the cell value to sanitize
        - warn: if True, print a warning when invalid characters are found
        """
        if not isinstance(value, str):
            return str(value)

        # Filter out invalid XML characters
        # Valid ranges: 0x09, 0x0A, 0x0D, 0x20-0xD7FF, 0xE000-0xFFFD, 0x10000-0x10FFFF
        sanitized = []
        invalid_chars = []
        for char in value:
            code = ord(char)
            if (code == 0x09 or code == 0x0A or code == 0x0D or
                (0x20 <= code <= 0xD7FF) or
                (0xE000 <= code <= 0xFFFD) or
                (0x10000 <= code <= 0x10FFFF)):
                sanitized.append(char)
            else:
                # Replace invalid character with its Unicode representation
                sanitized.append(f"[U+{code:04X}]")
                if warn:
                    invalid_chars.append(f"U+{code:04X}")

        if warn and invalid_chars:
            preview = value[:50] + "..." if len(value) > 50 else value
            print(f"Warning: Found invalid XML characters {invalid_chars} in cell: {preview}")

        return ''.join(sanitized)

    def _write_report(rows: list[dict[str, str]], path: Path) -> None:
        """Write a two-sheet Excel report (Results + Summary).

        Sheet: Results
        - Dataset
        - Row ID
        - Loaded DoclingDocument
        - Loaded DoclingDocument Error
        - Serialized IDocTags (mode, escape_mode, content) for all combinations
        - Serialized HTML
        - Serialized HTML Error

        Sheet: Summary
        - Metric, Count
        """
        # Build column list matching the actual columns generated
        cols = [
            "Dataset",
            "Row ID",
            "Loaded DoclingDocument",
            "Loaded DoclingDocument Error",
        ]

        # Add all combinations of mode, escape_mode, and content
        for mode in IDocTagsSerializationMode:
            for esc_mode in EscapeMode:
                for content in [True, False]:
                    cols.append(f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content})")
                    cols.append(f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content}) Error")

        cols.extend([
            "Serialized HTML",
            "Serialized HTML Error",
        ])

        # Ensure all rows have all columns and sanitize cell values
        norm_rows = []
        for r in rows:
            norm_rows.append({c: _sanitize_cell_value(r.get(c, ""), warn=True) for c in cols})

        # Build summary from normalized rows
        def _count_yes(key: str) -> int:
            return sum(1 for r in norm_rows if r.get(key, "") == "Yes")

        summary_rows = [
            {"Metric": "Total processed", "Count": len(norm_rows)},
            {"Metric": "Loaded DoclingDocument", "Count": _count_yes("Loaded DoclingDocument")},
        ]

        # Add summary rows for all combinations
        for mode in IDocTagsSerializationMode:
            for esc_mode in EscapeMode:
                for content in [True, False]:
                    col_name = f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content})"
                    summary_rows.append({"Metric": col_name, "Count": _count_yes(col_name)})

        summary_rows.append({"Metric": "Serialized HTML", "Count": _count_yes("Serialized HTML")})

        # Enforce .xlsx output; CSV export removed by request
        if path.suffix.lower() != ".xlsx":
            print(f"Report path '{path}' is not .xlsx; writing to .xlsx instead.")
            path = path.with_suffix(".xlsx")

        # Try pandas with ExcelWriter first
        try:
            import pandas as pd  # type: ignore
            from openpyxl import load_workbook  # type: ignore
            from openpyxl.styles import Alignment  # type: ignore

            df_results = pd.DataFrame(norm_rows, columns=cols)
            df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Count"])
            path.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df_results.to_excel(writer, sheet_name="Results", index=False)
                df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # Post-process with openpyxl to add styling
            wb = load_workbook(path)
            ws_results = wb["Results"]

            # Find error column indices (1-based)
            error_col_indices = []
            for idx, col_name in enumerate(cols, start=1):
                if "Error" in col_name:
                    error_col_indices.append(idx)

            # Apply styling to error columns
            for col_idx in error_col_indices:
                col_letter = ws_results.cell(row=1, column=col_idx).column_letter
                current_width = ws_results.column_dimensions[col_letter].width
                if current_width is None:
                    current_width = 8.43  # Default Excel column width
                ws_results.column_dimensions[col_letter].width = current_width * 3

                # Enable text wrapping for all cells in this column
                for row in range(1, ws_results.max_row + 1):
                    cell = ws_results.cell(row=row, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Enable text wrapping for all header cells
            for col_idx in range(1, len(cols) + 1):
                cell = ws_results.cell(row=1, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

            wb.save(path)
            print(f"Wrote report (Excel via pandas) to: {path}")
            return
        except Exception as exc_pd:
            print(f"pandas export failed ({exc_pd}); will try openpyxl.")

        # Try openpyxl second
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Alignment  # type: ignore

            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            # Results sheet
            ws_results = wb.active
            ws_results.title = "Results"
            ws_results.append(cols)
            for r in norm_rows:
                # norm_rows is already sanitized, but ensure column headers are also sanitized
                ws_results.append([r.get(c, "") for c in cols])

            # Find error column indices (1-based)
            error_col_indices = []
            for idx, col_name in enumerate(cols, start=1):
                if "Error" in col_name:
                    error_col_indices.append(idx)

            # Apply styling to error columns
            for col_idx in error_col_indices:
                col_letter = ws_results.cell(row=1, column=col_idx).column_letter
                current_width = ws_results.column_dimensions[col_letter].width
                if current_width is None:
                    current_width = 8.43  # Default Excel column width
                ws_results.column_dimensions[col_letter].width = current_width * 3

                # Enable text wrapping for all cells in this column
                for row in range(1, ws_results.max_row + 1):
                    cell = ws_results.cell(row=row, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Enable text wrapping for all header cells
            for col_idx in range(1, len(cols) + 1):
                cell = ws_results.cell(row=1, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

            # Summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.append(["Metric", "Count"])
            for r in summary_rows:
                ws_summary.append([_sanitize_cell_value(str(r["Metric"])), str(r["Count"])])
            wb.save(path)
            print(f"Wrote report (Excel via openpyxl) to: {path}")
            return
        except Exception as exc_xl:
            print(f"openpyxl export failed ({exc_xl}); report not written.")

    for idx, row in tqdm(enumerate(split), total=processed_total, ncols=128):
        if limit is not None and idx >= limit:
            break
        text = row.get("GroundTruthDocument", "")

        # Initialize per-row result
        row_result: dict[str, str] = {
            "Dataset": dataset_name,
            "Row ID": str(idx),
            "Loaded DoclingDocument": _yes(False),
            "Loaded DoclingDocument Error": "",
            "Serialized HTML": _yes(False),
            "Serialized HTML Error": "",
        }

        for mode in IDocTagsSerializationMode:
            for esc_mode in EscapeMode:
                for content in [True, False]:
                    row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content})"] = _yes(False)
                    row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content}) Error"] = ""

        try:
            doc = DoclingDocument.model_validate_json(text)
            page_images = [
                __ for __ in row["GroundTruthPageImages"]
            ]
            # page_images[0].show()
            row_result["Loaded DoclingDocument"] = _yes(True)
        except Exception as exc:
            errors.append(
                f"Parse error: {exc} for {dataset_name}/{dataset_subset}/{dataset_split} idx={idx}"
            )
            # Record failure outcome for this row
            row_result["Loaded DoclingDocument Error"] = str(exc)

            for mode in IDocTagsSerializationMode:
                for esc_mode in EscapeMode:
                    for content in [True, False]:
                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content})"] = _yes(False)
                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode.value}, content={content}) Error"] = "NA"

            results_rows.append(row_result)
            continue

        for i, __ in enumerate(page_images, start=1):
            doc.pages[i].image = ImageRef.from_pil(__, dpi=140)
            # png_path = pngs_dir / f"{idx}_{i}.png"
            # __.save(png_path)

        for mode in [IDocTagsSerializationMode.HUMAN_FRIENDLY, IDocTagsSerializationMode.LLM_FRIENDLY]:
            for esc_mode in [True, False]:
                for content in [True, False]:
                    try:
                        params_probe = IDocTagsParams()
                        params_probe.add_text_content = content
                        params_probe.mode = mode
                        params_probe.escape_mode = esc_mode
                        params_probe.pretty_indentation = "  " if mode==IDocTagsSerializationMode.HUMAN_FRIENDLY else None

                        iser_probe = IDocTagsDocSerializer(doc=doc, params=params_probe)
                        _ = iser_probe.serialize().text

                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content})"] = _yes(True)
                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content}) Error"] = ""

                    except Exception as exc_:
                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content})"] = _yes(False)
                        row_result[f"Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content}) Error"] = str(exc_)

        # Attempt HTML export (non-writing) to check serialization capability
        try:
            _ = doc.export_to_html(
                image_mode=ImageRefMode.EMBEDDED,
                split_page_view=True,
                include_annotations=True,
            )
            row_result["Serialized HTML"] = _yes(True)
        except Exception as html_exc:
            row_result["Serialized HTML"] = _yes(False)
            row_result["Serialized HTML Error"] = str(html_exc)

        # Append the result for this row
        results_rows.append(row_result)

    # Write report at the end
    try:
        _write_report(results_rows, report_path)
    except Exception as rep_exc:
        print(f"Failed to write report to {report_path}: {rep_exc}")

    # Print summary overview
    def _count_yes(rows: list[dict[str, str]], key: str) -> int:
        return sum(1 for r in rows if r.get(key, "") == "Yes")

    print("Overview summary:")
    print(f" - Total processed: {len(results_rows)}")
    print(f" - Loaded DoclingDocument: {_count_yes(results_rows, 'Loaded DoclingDocument')}")
    for mode in [IDocTagsSerializationMode.HUMAN_FRIENDLY, IDocTagsSerializationMode.LLM_FRIENDLY]:
        for esc_mode in [True, False]:
            for content in [True, False]:
                print(f" - Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content}): {_count_yes(results_rows, f'Serialized IDocTags ({mode.value}, escape_mode={esc_mode}, content={content})')}")
    print(f" - Serialized HTML: {_count_yes(results_rows, 'Serialized HTML')}")

    if errors:
        print("Errors:")
        for e in errors:
            print(" -", e)

    # Return 0 if no errors occurred, 1 otherwise
    return 0 if not errors else 1


def run_analyse(cfg: dict[str, Any]) -> int:
    """Analyse token lengths and special-token usage from IDocTags files.

    Config keys (with defaults):
    - tokenizer_name: str ("ibm-granite/granite-docling-258M")
    - input_glob: str ("./scratch/idoctags/*_with_h.idoctags")
    - pair_replace: dict with {"from": "_h", "to": ""} to locate paired files (optional)
    - show_plots: bool (True)
    - verbose: bool (False)
    """
    tokenizer_name = cfg.get("tokenizer_name", "ibm-granite/granite-docling-258M")
    input_glob_pat = cfg.get("input_glob", "./scratch/idoctags/*_with_h.idoctags")
    pair_replace = cfg.get("pair_replace", {"from": "_h", "to": ""})
    show_plots = bool(cfg.get("show_plots", True))
    verbose = bool(cfg.get("verbose", False))

    base_tokenizer: PreTrainedTokenizerBase = AutoTokenizer.from_pretrained(tokenizer_name)
    ext_tokenizer: PreTrainedTokenizerBase = AutoTokenizer.from_pretrained(tokenizer_name)
    ext_tokenizer = update_tokenizer(tokenizer=ext_tokenizer, verbose=verbose)

    filenames = sorted(glob.glob(input_glob_pat))
    if not filenames:
        print(f"No files matched pattern: {input_glob_pat}")
        return 1

    base_lengths: list[int] = []
    ext_lengths: list[int] = []
    special_counts: Counter[int] = Counter()

    # Map special tokens to IDs in the extended tokenizer
    special_tokens = IDocTagsVocabulary.get_special_tokens()
    special_token_ids = {ext_tokenizer.convert_tokens_to_ids(tok) for tok in special_tokens}

    for filename in tqdm(filenames, desc="Analyse", ncols=128):
        with open(filename, "r", encoding="utf-8") as fr:
            text_h = fr.read()

        paired_name: Optional[str] = None
        if isinstance(pair_replace, dict) and pair_replace.get("from") is not None:
            paired_name = filename.replace(str(pair_replace.get("from")), str(pair_replace.get("to", "")))

        text_plain = None
        if paired_name and os.path.exists(paired_name):
            with open(paired_name, "r", encoding="utf-8") as fr:
                text_plain = fr.read()

        # Tokenize and collect lengths
        base_tokens_h = base_tokenizer.encode(text_h, add_special_tokens=True)
        ext_tokens_h = ext_tokenizer.encode(text_h, add_special_tokens=True)
        base_lengths.append(len(base_tokens_h))
        ext_lengths.append(len(ext_tokens_h))

        # Special token usage on the extended-tokenized sequence of the HF/H text
        for tid in ext_tokens_h:
            if tid in special_token_ids:
                special_counts[tid] += 1

        # Optionally also include the paired non-human-friendly text lengths
        if text_plain is not None:
            _ = base_tokenizer.encode(text_plain, add_special_tokens=True)
            _e = ext_tokenizer.encode(text_plain, add_special_tokens=True)
            base_lengths.append(len(_))
            ext_lengths.append(len(_e))
            for tid in _e:
                if tid in special_token_ids:
                    special_counts[tid] += 1

    # Report summary
    print(f"Files analysed: {len(filenames)}")
    print(f"Sequences tokenized (including pairs when available): {len(ext_lengths)}")
    if ext_lengths:
        arr = np.asarray(ext_lengths)
        print(
            f"Extended tokenizer lengths — min: {arr.min()}, p50: {np.median(arr)}, p95: {np.percentile(arr,95)}, max: {arr.max()}, mean: {arr.mean():.1f}"
        )

    # Map special IDs back to tokens and show top-k
    if special_counts:
        id_to_token = {i: ext_tokenizer.convert_ids_to_tokens([i])[0] for i in special_counts.keys()}
        total_special = sum(special_counts.values())
        print("Special token usage (top 30):")
        for tid, cnt in special_counts.most_common(30):
            tok = id_to_token.get(tid, str(tid))
            pct = 100.0 * cnt / total_special if total_special else 0.0
            print(f" - {tok}: {cnt} ({pct:.2f}%)")

        if show_plots:
            # Bar chart of top-N special tokens
            items = special_counts.most_common(30)
            labels = [id_to_token.get(tid, str(tid)) for tid, _ in items]
            values = [v for _, v in items]
            plt.figure(figsize=(12, 5))
            plt.bar(labels, values, color="#4C78A8")
            plt.xticks(rotation=90)
            plt.title("Top Special Tokens (by frequency)")
            plt.tight_layout()
            plt.show()

    # Visualizations for lengths
    if show_plots and base_lengths and ext_lengths:
        plot_token_histograms(base_lengths, ext_lengths)
        plot_token_scatter_with_regression(base_lengths, ext_lengths)

    return 0


def plot_token_histograms(
    original_num_tokens: Sequence[int] | np.ndarray,
    optimal_num_tokens: Sequence[int] | np.ndarray,
    *,
    bins: int = 50,
    density: bool = False,
) -> None:
    """Plot overlapping histograms for original and optimal token counts.

    Parameters
    - original_num_tokens: sequence of counts before optimization
    - optimal_num_tokens: sequence of counts after optimization
    - bins: number of histogram bins
    - density: normalize histograms if True
    """
    x = np.asarray(original_num_tokens)
    y = np.asarray(optimal_num_tokens)

    plt.figure(figsize=(10, 6))
    plt.hist(
        x,
        bins=bins,
        alpha=0.5,
        label="Original",
        color="#4C78A8",
        edgecolor="black",
        density=density,
    )
    plt.hist(
        y,
        bins=bins,
        alpha=0.5,
        label="Optimal",
        color="#F58518",
        edgecolor="black",
        density=density,
    )
    plt.xlabel("Number of tokens")
    plt.ylabel("Density" if density else "Frequency")
    plt.title("Token Count Distributions")
    plt.legend()
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.show()


def plot_token_scatter_with_regression(
    original_num_tokens: Sequence[int] | np.ndarray,
    optimal_num_tokens: Sequence[int] | np.ndarray,
) -> None:
    """Scatter plot of original vs optimal tokens with a linear regression line.

    Parameters
    - original_num_tokens: x-values
    - optimal_num_tokens: y-values
    """
    x = np.asarray(original_num_tokens, dtype=float)
    y = np.asarray(optimal_num_tokens, dtype=float)

    if x.size == 0 or y.size == 0:
        print("No data to plot.")
        return

    # Linear regression with numpy polyfit (degree 1)
    slope, intercept = np.polyfit(x, y, 1)
    y_pred = slope * x + intercept
    # Compute R^2
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")

    # Prepare line for full x-range
    x_line = np.linspace(x.min(), x.max(), 100)
    y_line = slope * x_line + intercept

    plt.figure(figsize=(8, 8))
    plt.scatter(x, y, alpha=0.6, color="#4C78A8", label="Documents")
    plt.plot(x_line, y_line, color="#E45756", linewidth=2,
             label=f"y = {slope:.3f}x + {intercept:.3f} (R²={r2:.3f})")
    plt.xlabel("Original tokens")
    plt.ylabel("Optimal tokens")
    plt.title("Original vs Optimal Tokens with Linear Fit")
    plt.grid(alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.show()

def default_config(mode: str) -> dict[str, Any]:
    if mode == "dump":
        return {
            "dataset_name": "docling-project/doclaynet-set-a",
            "dataset_subset": "pdf_train",
            "dataset_split": "train",
            "output_dir": "./scratch/idoctags",
            "failed_dir": "./scratch/idoctags_failed",
            "report_path": "./scratch/idoctags_report.xlsx",
            "write_outputs": True,
            "limit": None,
            "variants": [
                {"add_content": False, "mode": "LLM_FRIENDLY", "suffix": "_without"},
                {"add_content": True, "mode": "LLM_FRIENDLY", "suffix": "_with"},
                {"add_content": True, "mode": "HUMAN_FRIENDLY", "suffix": "_with_h"},
            ],
        }
    elif mode == "analyse":
        return {
            "tokenizer_name": "ibm-granite/granite-docling-258M",
            "input_glob": "./scratch/idoctags/*_with_h.idoctags",
            "pair_replace": {"from": "_h", "to": ""},
            "show_plots": True,
            "verbose": False,
        }
    else:
        raise ValueError(f"Unknown mode for default config: {mode}")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Convert and analyse IDocTags data")
    parser.add_argument(
        "--mode",
        choices=["dump", "analyse"],
        required=True,
        help="Mode: dump dataset to idoctags or analyse token stats",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Path to JSON config. If omitted, a default config is created for the chosen mode.",
    )
    parser.add_argument(
        "--write-default-config",
        action="store_true",
        help="Only write the default config for the selected mode and exit.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="For dump mode: process only the first N items.",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    cfg_path: Optional[Path] = args.config
    cfg: dict[str, Any]

    if cfg_path is None:
        cfg = default_config(args.mode)
        out_name = Path(f"idoctags_{args.mode}_config.json")
        with open(out_name, "w", encoding="utf-8") as fw:
            json.dump(cfg, fw, indent=2)
        print(f"Wrote default config to: {out_name.resolve()}")
        if args.write_default_config:
            return 0
        # proceed using the freshly created config in-memory
    else:
        with open(cfg_path, "r", encoding="utf-8") as fr:
            cfg = json.load(fr)

    # Allow CLI --limit to override config for dump mode
    if args.mode == "dump" and args.limit is not None and args.limit > 0:
        cfg["limit"] = args.limit

    if args.mode == "dump":
        return run_dump(cfg)
    elif args.mode == "analyse":
        return run_analyse(cfg)
    else:
        raise AssertionError("unreachable")


if __name__ == "__main__":
    raise SystemExit(main())
